import json
import re
from datetime import date, datetime
from io import BytesIO

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import FileResponse
from openpyxl import load_workbook
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.config import ROOT_DIR
from app.database import get_db
from app.models.bom import BOMMaster
from app.models.material import MaterialMaster

router = APIRouter()

TEMPLATE_DIR = ROOT_DIR / "data" / "templates"
TEMPLATE_FILES = {
    "mps": TEMPLATE_DIR / "mps_template.xlsx",
    "bom": TEMPLATE_DIR / "bom_template.xlsx",
    "materials": TEMPLATE_DIR / "materials_template.xlsx",
}


@router.get("/templates/{name}")
def download_template(name: str):
    path = TEMPLATE_FILES.get(name)
    if not path or not path.exists():
        raise HTTPException(status_code=404, detail=f"模板 '{name}' 不存在")
    return FileResponse(
        path=str(path),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=path.name,
    )


@router.post("/bom/upload")
async def upload_bom(file: UploadFile = File(...), db: Session = Depends(get_db)) -> dict:
    content = await file.read()
    rows = _read_excel_rows(content)
    if not rows:
        raise HTTPException(status_code=400, detail="BOM 文件为空")

    headers = _normalize_headers(rows[0])
    required = {"parent_pn", "child_pn", "qty_per"}
    if not required.issubset(set(headers)):
        raise HTTPException(status_code=400, detail="BOM 缺少必填列: parent_pn, child_pn, qty_per")
    idx = {name: headers.index(name) for name in headers}

    upserted = 0
    for row in rows[1:]:
        parent = str(_cell(row, idx, "parent_pn") or "").strip()
        child = str(_cell(row, idx, "child_pn") or "").strip()
        qty_per_raw = _cell(row, idx, "qty_per")
        if not parent or not child or qty_per_raw in (None, ""):
            continue

        qty_per = float(qty_per_raw)
        supplier_name = str(_cell(row, idx, "supplier_name", default="") or "").strip()
        material_desc = str(_cell(row, idx, "material_desc", default="") or "").strip()
        material_type = str(_cell(row, idx, "material_type", default="") or "").strip()
        usage_uom = str(_cell(row, idx, "usage_uom", default="EA") or "EA").strip()
        level = int(float(_cell(row, idx, "level", default=1)))
        is_us = _parse_bool_like(_cell(row, idx, "is_us_material", default="N"))
        aml_text = str(_cell(row, idx, "aml", default="")).strip()
        aml = [part.strip() for part in aml_text.split(",") if part.strip()]

        existing = (
            db.query(BOMMaster)
            .filter(BOMMaster.parent_pn == parent, BOMMaster.child_pn == child)
            .one_or_none()
        )
        if existing:
            existing.supplier_name = supplier_name
            existing.material_desc = material_desc
            existing.material_type = material_type
            existing.qty_per = qty_per
            existing.usage_uom = usage_uom
            existing.level = level
            existing.is_us_material = is_us
            existing.aml_json = json.dumps(aml, ensure_ascii=False)
        else:
            db.add(
                BOMMaster(
                    parent_pn=parent,
                    child_pn=child,
                    supplier_name=supplier_name,
                    material_desc=material_desc,
                    material_type=material_type,
                    qty_per=qty_per,
                    usage_uom=usage_uom,
                    level=level,
                    is_us_material=is_us,
                    aml_json=json.dumps(aml, ensure_ascii=False),
                )
            )
        upserted += 1
    db.commit()
    return {"upserted": upserted}


@router.post("/materials/upload")
async def upload_materials(file: UploadFile = File(...), db: Session = Depends(get_db)) -> dict:
    content = await file.read()
    rows = _read_excel_rows(content)
    if not rows:
        raise HTTPException(status_code=400, detail="物料文件为空")

    headers = _normalize_headers(rows[0])
    required = {"material_pn", "lead_time_days"}
    if not required.issubset(set(headers)):
        raise HTTPException(status_code=400, detail="物料文件缺少必填列: material_pn, lead_time_days")
    idx = {name: headers.index(name) for name in headers}

    upserted = 0
    for row in rows[1:]:
        material_pn = _cell(row, idx, "material_pn")
        if not material_pn:
            continue
        lead_time_days = int(float(_cell(row, idx, "lead_time_days", default=0)))
        existing = db.query(MaterialMaster).filter(MaterialMaster.material_pn == material_pn).one_or_none()
        target = existing or MaterialMaster(material_pn=material_pn, lead_time_days=lead_time_days)

        target.supplier_name = str(_cell(row, idx, "supplier_name", default="") or "").strip()
        desc_val = _cell(row, idx, "description", default=None) or _cell(row, idx, "material_desc", default="")
        target.description = str(desc_val or "").strip()
        target.material_type = str(_cell(row, idx, "material_type", default="") or "").strip()
        target.lead_time_days = lead_time_days
        target.actual_delivery_date = _parse_date(_cell(row, idx, "actual_delivery_date"))
        target.shelf_life_expiry = _parse_date(_cell(row, idx, "shelf_life_expiry"))
        target.on_hand_inventory = float(_cell(row, idx, "on_hand_inventory", default=0) or 0)
        target.in_transit_inventory = float(_cell(row, idx, "in_transit_inventory", default=0) or 0)
        target.safety_stock = float(_cell(row, idx, "safety_stock", default=0) or 0)
        target.lot_size = float(_cell(row, idx, "lot_size", default=1) or 1)
        target.yield_rate = float(_cell(row, idx, "yield_rate", default=1) or 1)
        target.inventory_uom = str(_cell(row, idx, "inventory_uom", default="EA") or "EA").strip()
        if existing is None:
            db.add(target)
        upserted += 1

    db.commit()
    return {"upserted": upserted}


@router.get("/bom")
def get_bom_data(
    supplier_keyword: str | None = Query(default=None, description="按供应商名称模糊筛选"),
    db: Session = Depends(get_db),
) -> dict:
    query = db.query(BOMMaster)
    if supplier_keyword and supplier_keyword.strip():
        keyword = f"%{supplier_keyword.strip()}%"
        query = query.filter(func.lower(BOMMaster.supplier_name).like(func.lower(keyword)))
    rows = query.all()
    return {
        "items": [
            {
                "parent_pn": row.parent_pn,
                "child_pn": row.child_pn,
                "supplier_name": row.supplier_name or "",
                "material_desc": row.material_desc or "",
                "material_type": row.material_type or "",
                "qty_per": row.qty_per,
                "usage_uom": row.usage_uom or "EA",
                "level": row.level,
                "is_us_material": row.is_us_material,
                "aml": json.loads(row.aml_json or "[]"),
            }
            for row in rows
        ]
    }


@router.delete("/bom")
def delete_bom_data(db: Session = Depends(get_db)) -> dict:
    deleted = db.query(BOMMaster).delete(synchronize_session=False)
    db.commit()
    return {"deleted": deleted}


@router.get("/materials")
def get_material_data(
    supplier_keyword: str | None = Query(default=None, description="按供应商名称模糊筛选"),
    db: Session = Depends(get_db),
) -> dict:
    query = db.query(MaterialMaster)
    if supplier_keyword and supplier_keyword.strip():
        keyword = f"%{supplier_keyword.strip()}%"
        query = query.filter(func.lower(MaterialMaster.supplier_name).like(func.lower(keyword)))
    rows = query.all()
    return {
        "items": [
            {
                "material_pn": row.material_pn,
                "supplier_name": row.supplier_name or "",
                "description": row.description or "",
                "material_type": row.material_type or "",
                "lead_time_days": row.lead_time_days,
                "actual_delivery_date": row.actual_delivery_date.isoformat() if row.actual_delivery_date else None,
                "shelf_life_expiry": row.shelf_life_expiry.isoformat() if row.shelf_life_expiry else None,
                "on_hand_inventory": row.on_hand_inventory,
                "in_transit_inventory": row.in_transit_inventory,
                "safety_stock": row.safety_stock,
                "lot_size": row.lot_size,
                "yield_rate": row.yield_rate,
                "inventory_uom": row.inventory_uom or "EA",
            }
            for row in rows
        ]
    }


@router.delete("/materials")
def delete_material_data(db: Session = Depends(get_db)) -> dict:
    deleted = db.query(MaterialMaster).delete(synchronize_session=False)
    db.commit()
    return {"deleted": deleted}


def _read_excel_rows(content: bytes) -> list[tuple]:
    if not content:
        return []
    wb = load_workbook(filename=BytesIO(content), data_only=True)
    ws = wb.active
    return list(ws.iter_rows(values_only=True))


def _normalize_headers(row: tuple) -> list[str]:
    known = {
        "parent_pn",
        "child_pn",
        "supplier_name",
        "material_desc",
        "material_type",
        "qty_per",
        "usage_uom",
        "level",
        "is_us_material",
        "aml",
        "material_pn",
        "description",
        "inventory_uom",
        "lead_time_days",
        "actual_delivery_date",
        "shelf_life_expiry",
        "on_hand_inventory",
        "in_transit_inventory",
        "safety_stock",
        "lot_size",
        "yield_rate",
    }
    alias_map = {
        "parent_pn": {"上级料号", "父项料号", "父料号"},
        "child_pn": {"下级料号", "子项料号", "子料号"},
        "supplier_name": {"供应商", "供应商名称", "厂商", "vendor", "supplier"},
        "qty_per": {"单位用量", "单耗"},
        "level": {"bom层级", "bom 层级", "层级"},
        "is_us_material": {"美系标识", "美系物料", "是否美系"},
        "aml": {"可替代料", "替代料", "aml"},
        "material_pn": {"物料料号", "物料编码", "料号"},
        "material_desc": {"物料描述", "描述", "说明"},
        "description": {"description"},
        "material_type": {"物料类型", "类型", "分类"},
        "usage_uom": {"用量单位"},
        "inventory_uom": {"库存单位"},
        "lead_time_days": {"标准采购提前期天数", "提前期天数", "交期天数", "采购提前期", "采购提前期(天)"},
        "actual_delivery_date": {"实际到货日期"},
        "shelf_life_expiry": {"保质期截止", "保质期"},
        "on_hand_inventory": {"现有库存"},
        "in_transit_inventory": {"在途库存"},
        "safety_stock": {"安全库存"},
        "lot_size": {"批量", "最小批量"},
        "yield_rate": {"良率"},
    }

    sorted_known = sorted(known, key=len, reverse=True)

    normalized: list[str] = []
    for cell in row:
        raw = str(cell).strip().lower() if cell is not None else ""
        text = raw.replace("（", "(").replace("）", ")").replace("：", ":")

        matched = None
        for key in sorted_known:
            if key in text:
                matched = key
                break

        if matched is None:
            best_key = None
            best_len = 0
            for key, aliases in alias_map.items():
                for alias in aliases:
                    if alias in text and len(alias) > best_len:
                        best_key = key
                        best_len = len(alias)
            matched = best_key

        if matched is None:
            tokens = re.findall(r"[a-z_]+", text)
            matched = next((token for token in tokens if token in known), text)

        normalized.append(matched)

    return normalized


def _cell(row: tuple, idx: dict[str, int], key: str, default=None):
    pos = idx.get(key)
    if pos is None or pos >= len(row):
        return default
    value = row[pos]
    return default if value is None else value


def _parse_date(value) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        return datetime.strptime(value.strip(), "%Y-%m-%d").date()
    return None


def _parse_bool_like(value) -> bool:
    text = str(value or "").strip().lower()
    if not text:
        return False
    return text in {"y", "yes", "true", "1", "是", "美系"}
