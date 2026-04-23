from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
import re

from openpyxl import load_workbook


REQUIRED_COLUMNS = ("fg_pn", "qty", "due_date", "priority")


@dataclass(slots=True)
class ParsedOrder:
    fg_pn: str
    qty: int
    due_date: date
    priority: str


@dataclass(slots=True)
class ParsedConstraints:
    no_us_material: bool
    auto_grade: bool
    custom_notes: str
    deadline_override: date | None = None


@dataclass(slots=True)
class ParsedMPS:
    orders: list[ParsedOrder]
    constraints: ParsedConstraints


def parse_constraints_from_notes(notes: str | None) -> ParsedConstraints:
    """Phase 2 fallback parser; Phase 3 can replace this with Agent parsing."""
    text = (notes or "").strip()
    lowered = text.lower()
    no_us_material = ("禁用美系" in text) or ("no us material" in lowered) or ("no_us_material" in lowered)
    auto_grade = ("车规" in text) or ("auto-grade" in lowered) or ("auto_grade" in lowered)
    return ParsedConstraints(
        no_us_material=no_us_material,
        auto_grade=auto_grade,
        custom_notes=text,
        deadline_override=None,
    )


def parse_mps_excel(content: bytes, notes: str | None = None) -> ParsedMPS:
    workbook = load_workbook(filename=BytesIO(content), data_only=True)
    sheet = workbook.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("MPS 文件为空")

    header_row = _normalize_headers(rows[0])
    missing = [name for name in REQUIRED_COLUMNS if name not in header_row]
    if missing:
        raise ValueError(f"MPS 缺少必填列: {', '.join(missing)}")

    index = {name: header_row.index(name) for name in REQUIRED_COLUMNS}
    orders: list[ParsedOrder] = []
    for row in rows[1:]:
        if row is None:
            continue
        fg_raw = row[index["fg_pn"]]
        qty_raw = row[index["qty"]]
        due_raw = row[index["due_date"]]
        prio_raw = row[index["priority"]]

        if all(item in (None, "") for item in (fg_raw, qty_raw, due_raw, prio_raw)):
            continue

        fg_pn = str(fg_raw).strip()
        if not fg_pn:
            raise ValueError("fg_pn 不能为空")

        try:
            qty = int(float(qty_raw))
        except (TypeError, ValueError) as exc:
            raise ValueError(f"qty 非法: {qty_raw}") from exc
        if qty <= 0:
            raise ValueError("qty 必须大于 0")

        due_date = _normalize_date(due_raw)
        priority = _normalize_priority(prio_raw)

        orders.append(ParsedOrder(fg_pn=fg_pn, qty=qty, due_date=due_date, priority=priority))

    if not orders:
        raise ValueError("MPS 文件无有效订单")

    return ParsedMPS(orders=orders, constraints=parse_constraints_from_notes(notes))


def _normalize_priority(raw: object) -> str:
    if raw is None or str(raw).strip() == "":
        raise ValueError("priority 不能为空")
    s = str(raw).strip()
    sl = s.lower()
    if sl in {"high", "h"} or s in {"高", "高优先级", "高优"}:
        return "high"
    if sl in {"low", "l"} or s in {"低", "低优先级", "低优"}:
        return "low"
    raise ValueError(f"priority 非法: {raw}，请使用 high/low 或 高/低")


def _normalize_date(value: object) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return datetime.strptime(value.strip(), "%Y-%m-%d").date()
        except ValueError as exc:
            raise ValueError(f"due_date 非法: {value}，格式应为 YYYY-MM-DD") from exc
    raise ValueError(f"due_date 非法: {value}")


def _normalize_headers(row: tuple) -> list[str]:
    known = {"fg_pn", "qty", "due_date", "priority"}
    alias_map = {
        "fg_pn": {"成品料号", "产品料号", "成品编码"},
        "qty": {"需求数量", "数量", "需求量"},
        "due_date": {"需求交期", "交期", "需求日期", "交付日期"},
        "priority": {"优先级", "紧急程度"},
    }

    normalized: list[str] = []
    for cell in row:
        raw = str(cell).strip().lower() if cell is not None else ""
        text = raw.replace("（", "(").replace("）", ")").replace("：", ":")

        matched = None
        for key in known:
            if key in text:
                matched = key
                break

        if matched is None:
            for key, aliases in alias_map.items():
                if any(alias in text for alias in aliases):
                    matched = key
                    break

        if matched is None:
            # 兜底：提取英文 token，兼容 "字段名(field_name:xxx)" 形式。
            tokens = re.findall(r"[a-z_]+", text)
            matched = next((token for token in tokens if token in known), text)

        normalized.append(matched)

    return normalized
